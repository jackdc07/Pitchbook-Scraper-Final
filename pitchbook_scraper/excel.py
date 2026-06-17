"""Write scraped companies to a formatted .xlsx spreadsheet."""
from __future__ import annotations

from pathlib import Path
from typing import Iterable

from .models import Company

# Column header -> how to pull the value from a Company.
_COLUMNS: list[tuple[str, callable]] = [
    ("Company", lambda c: c.name or ""),
    ("Website", lambda c: c.website or ""),
    ("Keywords", lambda c: c.keywords or ""),
    ("Employees", lambda c: c.employees or ""),
    ("Last Round Type", lambda c: c.last_round_type or ""),
    ("Last Round Amount", lambda c: c.last_round_amount or ""),
    ("Last Round Date", lambda c: c.last_round_date or ""),
    ("Total Raised", lambda c: c.total_raised or ""),
    ("Most Recent Revenue", lambda c: c.most_recent_revenue or ""),
    ("Revenue Date", lambda c: c.revenue_date or ""),
    ("Acquired", lambda c: "" if c.acquired is None else ("Yes" if c.acquired else "No")),
    ("Acquirer", lambda c: c.acquirer or ""),
    ("Acquisition Date", lambda c: c.acquisition_date or ""),
    ("Team Size", lambda c: c.team_size),
    ("Team Names", lambda c: "; ".join(c.team_names)),
    ("Team Emails", lambda c: "; ".join(c.team_emails)),
    ("Team", lambda c: "; ".join(f"{m.name} ({m.title})" if m.title else m.name for m in c.team)),
    ("Primary Offices", lambda c: "; ".join(c.primary_offices)),
    ("Description", lambda c: c.description or ""),
    ("Source File", lambda c: c.source_file or ""),
]


def write_xlsx(companies: Iterable[Company], path: str | Path) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise ImportError(
            "openpyxl is required for Excel output. Install with: pip install openpyxl"
        ) from exc

    companies = list(companies)
    path = Path(path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Companies"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")

    # Header row.
    for col, (name, _) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    # Data rows.
    for r, company in enumerate(companies, start=2):
        for col, (_, getter) in enumerate(_COLUMNS, start=1):
            ws.cell(row=r, column=col, value=getter(company))

    # Reasonable column widths (cap wide text columns).
    for col, (name, _) in enumerate(_COLUMNS, start=1):
        letter = get_column_letter(col)
        if name in ("Description", "Team", "Primary Offices"):
            ws.column_dimensions[letter].width = 50
        elif name in ("Keywords", "Team Names", "Team Emails"):
            ws.column_dimensions[letter].width = 32
        else:
            ws.column_dimensions[letter].width = max(14, len(name) + 2)

    wb.save(str(path))
    return path
